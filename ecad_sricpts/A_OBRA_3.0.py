import os
import re
import time
import warnings
import logging

import PyPDF2
import pandas as pd
from openpyxl import Workbook, load_workbook
from plyer import notification

# === CONFIGURAÇÃO DE LOG ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# === SUPRIMIR AVISOS ===
warnings.filterwarnings("ignore", category=UserWarning, module='pdfminer')

# ==== 🗂️ Caminhos ====
folder_path     = r'C:\Users\guilherme.cristo\Documents\ECAD\s_pdf_organizados'
output_folder   = r'C:\Users\guilherme.cristo\Documents\ECAD\s_tabelas\obras'
arquivo_modelo  = r'C:\Users\guilherme.cristo\Documents\ECAD\en_modelo\Modelo_Valuation_vfinal_4.0.xlsm'
pasta_origem    = output_folder
pasta_destino   = r'C:\Users\guilherme.cristo\Documents\ECAD\s_tabelas\compiladas'

# ==== 📄 Funções ====

def extract_text_from_pdf(pdf_path):
    text = ""
    with open(pdf_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text += page.extract_text() or ""
    return text

def extract_data_referente(text):
    m = re.search(r'\b[A-ZÇ]{3,9}/\d{4}\b', text)
    return m.group(0) if m else None

def process_text(text):
    data = []
    lines = text.split('\n')
    collect = False
    total_from_pdf = 0.0
    capture_next_line = False
    data_referente = extract_data_referente(text)

    for line in lines:
        if 'OBRA' in line:
            collect = True

        if capture_next_line:
            m = re.search(r'([\d.,]+)', line)
            if m:
                s = m.group(1).replace('.', '').replace(',', '.')
                if s.count('.') > 1:
                    parts = s.split('.')
                    s = ''.join(parts[:-1]) + '.' + parts[-1]
                try:
                    total_from_pdf = float(s)
                except ValueError:
                    logging.error(f"Erro ao converter '{s}' em float.")
                capture_next_line = False

        if 'TOTAL GERAL' in line:
            capture_next_line = True

        if collect and re.match(r'^\d{2,}', line):
            parts = line.split()
            codigo = parts[0]
            try:
                idx = next(i for i,p in enumerate(parts)
                           if re.match(r'^\d{1,3}(?:\.\d{3})*,\d{2}|\d{9}$', p))
                nome = ' '.join(parts[1:idx])
                if nome:
                    rateio_raw = parts[idx]
                    rateio_tratado = rateio_raw.replace('.', '').replace(',', '.')
                    try:
                        rateio_float = float(rateio_tratado)
                        data.append([codigo, nome, rateio_float, data_referente])
                    except ValueError:
                            # Ignora a linha problemática sem adicionar nada
                            logging.warning(f"🔴 Linha ignorada: rateio inválido '{rateio_raw}' no arquivo. Conteúdo: '{line}'")
                            continue
            except StopIteration:
                continue

    return data, total_from_pdf

def save_to_excel(data, output_path, pdf_filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados ECAD"
    ws.append(['Nome Arquivo','Código ECAD','Nome Obra','Rateio','Data'])

    total_rateio = 0.0
    for row in data:
        ws.append([pdf_filename] + row)
        total_rateio += row[2]

    wb.save(output_path)
    return total_rateio


def formatar_dataframe_obras(df: pd.DataFrame) -> pd.DataFrame:
    df["Nome Arquivo"] = df["Nome Arquivo"].astype(str)
    df["Nome Obra"] = df["Nome Obra"].astype(str)
    df["Código ECAD"] = pd.to_numeric(df["Código ECAD"], errors="coerce").fillna(0).astype(int)

    # Função segura para tratar valores com ponto ou vírgula decimal
    def limpar_valor_rateio(val):
        if pd.isna(val) or str(val).strip() == '---':
            return 0.0
        val = str(val).strip()
        if '.' in val and ',' not in val:
            try:
                return float(val)
            except ValueError:
                return 0.0
        val = val.replace('.', '').replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return 0.0

    df["Rateio"] = df["Rateio"].apply(limpar_valor_rateio)

    def converter_data(texto):
        meses = {
            "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "ABRIL": "04",
            "MAIO": "05", "JUNHO": "06", "JULHO": "07", "AGOSTO": "08",
            "SETEMBRO": "09", "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12"
        }
        if isinstance(texto, str):
            texto = texto.upper()
            for nome_mes, num_mes in meses.items():
                if nome_mes in texto:
                    ano_match = re.search(r'\d{4}', texto)
                    if ano_match:
                        return pd.to_datetime(f"01/{num_mes}/{ano_match.group(0)}", dayfirst=True, errors='coerce')
        return pd.NaT

    df["Data"] = df["Data"].apply(converter_data)

    return df


def main():
    inicio = time.time()

    # cria pasta de saída, se necessário
    os.makedirs(output_folder, exist_ok=True)

    # ==== 📥 Processar PDFs ====
    for fname in os.listdir(folder_path):
        if not fname.lower().endswith('.pdf'):
            continue

        pdf_path    = os.path.join(folder_path, fname)
        out_path    = os.path.join(output_folder, fname.replace('.pdf', '.xlsx'))

        logging.info(f"Processando {fname}...")
        text, data = extract_text_from_pdf(pdf_path), None
        data, total_pdf = process_text(text)
        total_rateio = save_to_excel(data, out_path, fname)

        tr = round(total_rateio, 2)
        tp = round(total_pdf, 2)
        if tr == tp:
            logging.info(f"Valores iguais: {tr:.2f}")
        else:
            logging.warning(f"Valores diferentes. Calculado: {tr:.2f}, PDF: {tp:.2f}")

        logging.info(f"Dados salvos em {out_path}")

    # ==== 📊 Compilar arquivos Excel ====
    excels = [f for f in os.listdir(pasta_origem) if f.lower().endswith('.xlsx')]
    dfs = []
    for f in excels:
        path = os.path.join(pasta_origem, f)
        try:
            dfs.append(pd.read_excel(path, engine='openpyxl'))
        except Exception as e:
            logging.error(f"Erro ao ler {f}: {e}")

    dfs = [df.dropna(axis=1, how='all') for df in dfs]
    
    
    df_comp = pd.concat(dfs, ignore_index=True)

    # aplicar formatação robusta (Rateio, Data, tipos corretos)
    df_comp = formatar_dataframe_obras(df_comp)

    # filtragens finais (remoção de ruídos)
    df_comp = df_comp[~df_comp['Código ECAD'].astype(str).str.contains(r'ª|º|°', na=False)]
    df_comp = df_comp[~df_comp['Nome Obra'].str.contains(r'\b\d{2}/\d{4}\b', na=False)]

    out_comp = os.path.join(pasta_destino, 'tabela_compilada_Obras.xlsx')
    df_comp.to_excel(out_comp, index=False, engine='openpyxl')
    logging.info(f"Arquivo compilado salvo em: {out_comp}")

    # ==== 📤 Inserir dados no XLSM ====
    while True:
        try:
            with open(arquivo_modelo, 'rb'):
                break
        except PermissionError:
            logging.warning("⚠️ Arquivo em uso, aguardando liberação...")
            time.sleep(1)

    wb = load_workbook(arquivo_modelo, keep_vba=True)
    ws = wb["bs_ECAD_Obras"]

    # escrita dos cabeçalhos
    for ci, col in enumerate(df_comp.columns, start=1):
        ws.cell(row=1, column=ci, value=col)
    # escrita dos dados
    for ri, row in enumerate(df_comp.itertuples(index=False, name=None), start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)

    wb.save(arquivo_modelo)
    fim = time.time()
    logging.info(f"✅ Exportação concluída em {fim - inicio:.2f} segundos.")

    notification.notify(
        title='PROCESSAMENTO DE OBRAS',
        message='Processamento finalizado!',
        timeout=120
    )

if __name__ == "__main__":
    main()
