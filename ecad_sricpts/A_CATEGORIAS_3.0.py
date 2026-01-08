import os
import time
import re
import warnings
import logging
import zipfile

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from plyer import notification

# === CONFIGURAÇÃO DE LOG ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
warnings.filterwarnings("ignore", category=UserWarning, module="pdfminer")
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# === CAMINHOS ===
PASTA_PDFS = r"C:\Users\guilherme.cristo\Documents\ECAD\s_pdf_organizados"
PASTA_EXCEL = r"C:\Users\guilherme.cristo\Documents\ECAD\s_tabelas\categorias"
ARQUIVO_COMPILADO = r"C:\Users\guilherme.cristo\Documents\ECAD\s_tabelas\compiladas\tabela_compilada_categorias.xlsx"
ARQUIVO_XLSM = r"C:\Users\guilherme.cristo\Documents\ECAD\en_modelo\Modelo_Valuation_vfinal_4.0.xlsm"
ABA_XLSM = "bs_ECAD_categorias"

# === EXPRESSÕES REGULARES ===
date_pattern = re.compile(r'\b[A-ZÇ]{3,9}/\d{4}\b')
numeric_pattern = re.compile(r'^\d{1,3}(?:\.\d{3})*,\d{2}$')
year_pattern = re.compile(r'^\d{4}$')

def extract_data_referente(text):
    match = date_pattern.search(text)
    return match.group(0) if match else None

def process_pdf(pdf_path):
    filename = os.path.basename(pdf_path)
    logging.info(f"🔍 Processando: {filename}")

    start_page = end_page = None
    data_referente = None
    page_texts = []

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            page_texts.append(text)

            # atualiza data referente se encontrada
            if (d := extract_data_referente(text)):
                data_referente = d

            if start_page is None and "POR CATEGORIA" in text:
                start_page = i
            if start_page and " POR RUBRICA" in text:
                end_page = i
                break

    if start_page is None or end_page is None:
        logging.error(f"❌ Tabela não encontrada em: {filename}")
        return

    # recorta o texto entre início e fim
    segments = []
    for i in range(start_page, end_page + 1):
        txt = page_texts[i - 1]
        if i == start_page:
            idx = txt.find("POR CATEGORIA")
            txt = txt[idx:] if idx != -1 else txt
        if i == end_page:
            idx = txt.find("POR RUBRICA")
            txt = txt[: idx + len("POR RUBRICA")] if idx != -1 else txt
        segments.append(txt)

    combined = "\n".join(segments).splitlines()

    header = [
        "CATEGORIA", "DISTRIBUIÇÃO", "LIBERAÇÃO CRÉD. RETIDO",
        "LIBERAÇÃO DE PENDENTE", "LIBERAÇÃO DE PARÂMETRO",
        "AJUSTES", "TOTAL GERAL", "DATA REFERENTE"
    ]
    data = []
    exclude = False

    for line in combined:
        if line.startswith("EXEC. - NÚM. DE EXECUÇÕES"):
            exclude = True
        if line.startswith("OBRA RUBRICA PERÍODO RENDIMENTO % RATEIO CORREÇÃO EXEC (OC"):
            exclude = False
            continue
        if exclude or not line.strip() or line.startswith(("POR CATEGORIA", "TOTAL")):
            continue

        parts = line.split()
        rubrica_parts, valores = [], []
        for p in parts:
            if numeric_pattern.match(p) or p == "---":
                valores.append(p)
            else:
                rubrica_parts.append(p)

        nome_rubrica = " ".join(rubrica_parts)
        while len(valores) < len(header) - 2:
            valores.insert(0, "---")
        valores.append(data_referente)
        data.append([nome_rubrica] + valores)

    df = pd.DataFrame(data, columns=header)
    nome_excel = f"tabela_extraida_{os.path.splitext(filename)[0]}.xlsx"
    caminho_excel = os.path.join(PASTA_EXCEL, nome_excel)
    df.to_excel(caminho_excel, index=False)
    logging.info(f"✅ Exportado para: {caminho_excel}")


def compilar_excels(pasta_excel: str) -> pd.DataFrame:
    arquivos = [f for f in os.listdir(pasta_excel) if f.lower().endswith(".xlsx")]
    logging.info(f"📄 {len(arquivos)} arquivos encontrados.")

    dfs = []
    for arquivo in arquivos:
        caminho = os.path.join(pasta_excel, arquivo)
        try:
            df = pd.read_excel(caminho)
            dfs.append(df)
        except Exception as e:
            logging.error(f"Erro ao ler {arquivo}: {e}")

    if not dfs:
        logging.warning("⚠️ Nenhum arquivo carregado.")
        return pd.DataFrame()

    compilado = pd.concat(dfs, ignore_index=True)

    if "TOTAL GERAL" in compilado.columns:
        compilado = compilado[compilado["TOTAL GERAL"] != '---']

    return compilado


def salvar_compilado(df: pd.DataFrame, caminho_saida: str):
    if df.empty:
        logging.warning("⚠️ DataFrame vazio, nada foi salvo.")
        return
    df.to_excel(caminho_saida, index=False)
    logging.info(f"📁 Compilado salvo em: {caminho_saida}")
    

def formatar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    colunas_numericas = [
        "DISTRIBUIÇÃO",
        "LIBERAÇÃO CRÉD. RETIDO",
        "LIBERAÇÃO DE PENDENTE",
        "LIBERAÇÃO DE PARÂMETRO",
        "AJUSTES",
        "TOTAL GERAL"
    ]

    # Função robusta para tratar todos os formatos numéricos
    def limpar_valor(val):
        if pd.isna(val) or str(val).strip() == '---':
            return 0.0
        val = str(val).strip()
        # Se já estiver com ponto decimal e sem vírgula
        if '.' in val and ',' not in val:
            try:
                return float(val)
            except ValueError:
                return 0.0
        # Caso contrário, aplica lógica para formato brasileiro
        val = val.replace('.', '').replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return 0.0

    # Aplicar a limpeza para cada coluna numérica
    for col in colunas_numericas:
        if col in df.columns:
            df[col] = df[col].apply(limpar_valor)

    # Converter DATA REFERENTE → datetime (01/mm/aaaa)
    if "DATA REFERENTE" in df.columns:
        def converter_data(texto):
            meses = {
                "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "ABRIL": "04",
                "MAIO": "05", "JUNHO": "06", "JULHO": "07", "AGOSTO": "08",
                "SETEMBRO": "09", "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12"
            }
            if isinstance(texto, str):
                texto = texto.upper()
                for nome_mes, num_mes in meses.items():
                    if nome_mes in texto:
                        match = re.search(r'\d{4}', texto)
                        if match:
                            return pd.to_datetime(f"01/{num_mes}/{match.group(0)}", dayfirst=True, errors='coerce')
            return pd.NaT

        df["DATA REFERENTE"] = df["DATA REFERENTE"].apply(converter_data)

    # Garantir que Período e Rubrica_Modelo sejam texto
    for col in ["Período", "Rubrica_Modelo"]:
        if col in df.columns:
            df[col] = df[col].astype(str)

    return df

def main():
    inicio = time.time()

    # 1. PROCESSAR TODOS OS PDFs
    for arquivo in os.listdir(PASTA_PDFS):
        if arquivo.lower().endswith(".pdf"):
            process_pdf(os.path.join(PASTA_PDFS, arquivo))


    # 2. COMPILAR EXCELS
    df_compilado = compilar_excels(PASTA_EXCEL)

    # 2.1 AJUSTAR FORMATOS
    df_compilado = formatar_dataframe(df_compilado)

    # 2.2 SALVAR
    salvar_compilado(df_compilado, ARQUIVO_COMPILADO)

    # 3. INSERIR NO XLSM
    while True:
        try:
            wb = load_workbook(ARQUIVO_XLSM, keep_vba=True)
            break
        except(PermissionError, zipfile.BadZipFile) as e:
            logging.warning(
                f"⚠️ Arquivo XLSM indisponível:({e.__class__.__name__}),aguardando 10 s para tentar de novo..."  
            )
            time.sleep(10)
    ws = wb[ABA_XLSM]
    for ci, nome in enumerate(df_compilado.columns, start=1):
        ws.cell(row=1, column=ci, value=nome)
    for ri, row in enumerate(df_compilado.itertuples(index=False, name=None), start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    wb.save(ARQUIVO_XLSM)
    logging.info(f"📥 Dados inseridos na aba '{ABA_XLSM}' do XLSM.")

    duracao = time.time() - inicio
    logging.info(f"⏱️ Tempo total de execução: {duracao:.2f} s")

    notification.notify(
        title='PROCESSAMENTO DE CATEGORIAS',
        message='Processamento finalizado!',
        timeout=10
    )

if __name__ == "__main__":
    main()
