import os
import time
import re
import warnings
import logging

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from plyer import notification

# === CONFIGURAÇÃO DE LOG ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
warnings.filterwarnings("ignore", category=UserWarning, module='pdfminer')
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# === CAMINHOS ===
PASTA_PDFS = r"C:\Users\guilherme.cristo\Documents\ECAD\s_pdf_organizados"
PASTA_EXCEL = r"C:\Users\guilherme.cristo\Documents\ECAD\s_tabelas\rubricas"
ARQUIVO_COMPILADO = r"C:\Users\guilherme.cristo\Documents\ECAD\s_tabelas\compiladas\tabela_compilada_rubricas.xlsx"
BASE_RUBRICAS = r"C:\Users\guilherme.cristo\Documents\ECAD\bs_rubricas\Base_Rubrica_Original.xlsx"
ARQUIVO_XLSM = r"C:\Users\guilherme.cristo\Documents\ECAD\en_modelo\Modelo_Valuation_vfinal_4.0.xlsm"

# === EXPRESSÕES REGULARES ===
date_pattern = re.compile(r'\b[A-ZÇ]{3,9}/\d{4}\b')
numeric_pattern = re.compile(r'^\d{1,3}(?:\.\d{3})*,\d{2}$')
year_pattern = re.compile(r'^\d{4}$')

def extract_data_referente(text):
    match = date_pattern.search(text)
    return match.group(0) if match else None

def process_pdf(pdf_path):
    filename = os.path.basename(pdf_path)
    logging.info(f"🔍 Iniciando processamento de: {filename}")

    start_page = end_page = None
    data_referente = None
    page_texts = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            page_texts.append(text)

            if (curr := extract_data_referente(text)):
                data_referente = curr

            if start_page is None and "POR RUBRICA" in text:
                start_page = page_number
            if start_page is not None and "TOTAL DO TITULAR" in text:
                end_page = page_number
                break

    if start_page is None or end_page is None:
        logging.error(f"❌ Tabela não encontrada em: {filename}")
        return

    segments = []
    for page_number in range(start_page, end_page + 1):
        txt = page_texts[page_number - 1]
        if page_number == start_page:
            idx = txt.find("POR RUBRICA")
            txt = txt[idx:] if idx != -1 else txt
        if page_number == end_page:
            idx = txt.find("TOTAL DO TITULAR")
            txt = txt[: idx + len("TOTAL DO TITULAR")] if idx != -1 else txt
        segments.append(txt)

    combined = "\n".join(segments).splitlines()

    header = [
        "RUBRICA", "DISTRIBUIÇÃO", "LIBERAÇÃO CRÉD. RETIDO",
        "LIBERAÇÃO DE PENDENTE", "LIBERAÇÃO DE PARÂMETRO",
        "AJUSTES", "TOTAL GERAL", "DATA REFERENTE"
    ]
    data = []
    exclude = False

    for line in combined:
        if line.startswith("EXEC. - NÚM. DE EXECUÇÕES"):
            exclude = True
        if line.startswith("OBRA RUBRICA PERÍODO RENDIMENTO % RATEIO CORREÇÃO EXEC (OC"):
            exclude = False
            continue
        if exclude or not line.strip() or line.startswith(("POR RUBRICA", "TOTAL")):
            continue

        parts = line.split()
        rubrica_parts, valores = [], []

        for part in parts:
            if numeric_pattern.match(part) or part == "---":
                valores.append(part)
            else:
                rubrica_parts.append(part)

        rubrica_name = " ".join(rubrica_parts)
        while len(valores) < len(header) - 2:
            valores.insert(0, "---")
        valores.append(data_referente)
        data.append([rubrica_name] + valores)

    df = pd.DataFrame(data, columns=header)
    nome_arquivo = f"tabela_extraida_{os.path.splitext(filename)[0]}.xlsx"
    caminho_excel = os.path.join(PASTA_EXCEL, nome_arquivo)
    df.to_excel(caminho_excel, index=False)
    logging.info(f"✅ Exportado para: {caminho_excel}")

def extract_period(text):
    match = re.search(r'\b\d{2}/\d{4}(?: A \d{2}/\d{4})?\b', text)
    return match.group() if match else None

def formatar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    colunas_numericas = [
        "DISTRIBUIÇÃO",
        "LIBERAÇÃO CRÉD. RETIDO",
        "LIBERAÇÃO DE PENDENTE",
        "LIBERAÇÃO DE PARÂMETRO",
        "AJUSTES",
        "TOTAL GERAL"
    ]

    # Função robusta para tratar todos os formatos numéricos
    def limpar_valor(val):
        if pd.isna(val) or str(val).strip() == '---':
            return 0.0
        val = str(val).strip()
        # Se já estiver com ponto decimal e sem vírgula
        if '.' in val and ',' not in val:
            try:
                return float(val)
            except ValueError:
                return 0.0
        # Caso contrário, aplica lógica para formato brasileiro
        val = val.replace('.', '').replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return 0.0

    # Aplicar a limpeza para cada coluna numérica
    for col in colunas_numericas:
        if col in df.columns:
            df[col] = df[col].apply(limpar_valor)

    # Converter DATA REFERENTE → datetime (01/mm/aaaa)
    if "DATA REFERENTE" in df.columns:
        def converter_data(texto):
            meses = {
                "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "ABRIL": "04",
                "MAIO": "05", "JUNHO": "06", "JULHO": "07", "AGOSTO": "08",
                "SETEMBRO": "09", "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12"
            }
            if isinstance(texto, str):
                texto = texto.upper()
                for nome_mes, num_mes in meses.items():
                    if nome_mes in texto:
                        match = re.search(r'\d{4}', texto)
                        if match:
                            return pd.to_datetime(f"01/{num_mes}/{match.group(0)}", dayfirst=True, errors='coerce')
            return pd.NaT

        df["DATA REFERENTE"] = df["DATA REFERENTE"].apply(converter_data)

    # Garantir que Período e Rubrica_Modelo sejam texto
    for col in ["Período", "Rubrica_Modelo"]:
        if col in df.columns:
            df[col] = df[col].astype(str)

    return df

def main():
    inicio = time.time()

    # 1. PROCESSAR PDFs
    for arquivo in os.listdir(PASTA_PDFS):
        if arquivo.lower().endswith(".pdf"):
            process_pdf(os.path.join(PASTA_PDFS, arquivo))

    # 2. COMPILAR OS EXCELS EXTRAÍDOS
    excels = [f for f in os.listdir(PASTA_EXCEL) if f.lower().endswith(".xlsx")]
    dfs = [pd.read_excel(os.path.join(PASTA_EXCEL, f)) for f in excels]
    df_compilado = pd.concat(dfs, ignore_index=True)
    df_compilado = df_compilado[df_compilado['TOTAL GERAL'] != '---']
    df_compilado['Período'] = df_compilado['RUBRICA'].apply(extract_period)
    df_compilado['RUBRICA'] = df_compilado['RUBRICA'].str.replace(
        r'\b\d{2}/\d{4}(?: A \d{2}/\d{4})?\b', '', regex=True
    ).str.strip()

    # 3. ADICIONAR RUBRICA MODELO
    base_rubricas = pd.read_excel(BASE_RUBRICAS, sheet_name=0)
    mapa = base_rubricas.set_index('Descrição')['Rubrica MODELO'].to_dict()
    df_compilado['Rubrica_Modelo'] = df_compilado['RUBRICA'].map(mapa)

    # 3.1 AJUSTAR FORMATOS PARA EXCEL
    df_compilado = formatar_dataframe(df_compilado)

    # 4. SALVAR COMPILADO FINAL
    df_compilado.to_excel(ARQUIVO_COMPILADO, index=False)
    logging.info(f"📁 Arquivo compilado salvo em: {ARQUIVO_COMPILADO}")

    # 5. INSERIR DADOS NO XLSM
    while True:
        try:
            with open(ARQUIVO_XLSM, 'rb'):
                break
        except PermissionError:
            logging.warning("⚠️ Arquivo XLSM em uso, aguardando liberação...")
            time.sleep(1)

    wb = load_workbook(ARQUIVO_XLSM, keep_vba=True)
    ws = wb["bs_ECAD_Rubrica"]
    for col_idx, col_name in enumerate(df_compilado.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_compilado.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(ARQUIVO_XLSM)
    logging.info(f"📥 Dados inseridos na planilha XLSM: {ARQUIVO_XLSM}")

    # 6. NOTIFICAÇÃO E LOG DE CONCLUSÃO
    duracao = time.time() - inicio
    logging.info(f"⏱️ Tempo total de execução: {duracao:.2f} segundos.")

    notification.notify(
        title='PROCESSAMENTO DE RUBRICAS',
        message='Processamento finalizado!',
        timeout=10
    )

if __name__ == "__main__":
    main()
